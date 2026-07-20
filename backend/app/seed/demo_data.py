"""
Seed do modo demo — M5.1. Reset automático/manual — ADR "reset periódico demo".

Uso via CLI (conexão própria, fora do pool do app):

    python -m app.seed.demo_data

Uso programático (dentro do app — worker periódico e endpoint de reset manual,
ver app/seed/demo_reset_service.py): `_build_demo_data(conn, demo_user_id, ...)`
recebe uma conexão já aberta (do pool admin) em vez de abrir a sua própria —
é o núcleo reaproveitado por ambos os caminhos.

Dados criados por empresa:
  - 1 usuário demo (username=demo, admin nas 3 empresas)
  - 2 usuários de equipe extras (Operador/Visualizador) — mostram o sistema
    de papéis populado sem precisar que um visitante crie usuários pra ver
  - 4 pastas raiz (Fiscal, RH, Bancário, Contratos)
  - ~17 documentos distribuídos (ocr=done maioria, 2 failed, 3 na lixeira)
  - ~3 favoritos (âncoras) e 2 links de compartilhamento (1 com senha)
  - activity_log simulado nos últimos 7 dias
"""

from __future__ import annotations

import asyncio
import bcrypt
import hashlib
import io
import random
import secrets
import uuid
import zipfile
from datetime import datetime, timedelta, timezone

import asyncpg
import httpx
from PIL import Image, ImageDraw
from openpyxl import Workbook

from app.config import settings
from app.services import storage_service


# ---------------------------------------------------------------------------
# Dados fictícios
# ---------------------------------------------------------------------------

COMPANIES = [
    ("Comércio Alfa Modelo Ltda",   "CNPJ: 12.345.678/0001-90"),
    ("Serviços Beta Referência SA", "CNPJ: 23.456.789/0001-01"),
    ("Indústria Gama Exemplo Ltda", "CNPJ: 34.567.890/0001-12"),
]

FOLDERS = ["Fiscal", "RH", "Bancário", "Contratos"]

# (name_template, ext, folder_index)
DOCUMENTS = [
    ("NF-e_Fornecedor_{n}.xml",     "xml",  0),
    ("SPED_Fiscal_{n}.txt",          "txt",  0),
    ("DCTF_Mensal_{n}.pdf",          "pdf",  0),
    ("Extrato_Banco_{n}.pdf",        "pdf",  2),
    ("Conciliação_{n}.xlsx",         "xlsx", 2),
    ("Boleto_{n}.pdf",               "pdf",  2),
    ("Contrato_Prestação_{n}.docx",  "docx", 3),
    ("Procuração_{n}.pdf",           "pdf",  3),
    ("Aditivo_{n}.docx",             "docx", 3),
    ("Holerite_{n}.pdf",             "pdf",  1),
    ("Admissão_{n}.pdf",             "pdf",  1),
    ("Rescisão_{n}.pdf",             "pdf",  1),
    ("Planilha_RH_{n}.xlsx",         "xlsx", 1),
    ("Imagem_Doc_{n}.jpg",           "jpg",  0),
    ("Relatorio_{n}.pdf",            "pdf",  2),
]

OCR_TEXTS = [
    "Nota fiscal eletrônica emitida em conformidade com a legislação vigente. "
    "CNPJ emitente: 12.345.678/0001-90. Valor total: R$ 1.250,00.",
    "Relatório de conciliação bancária referente ao período. "
    "Saldo inicial: R$ 50.000,00. Saldo final: R$ 48.750,00.",
    "Contrato de prestação de serviços firmado entre as partes. "
    "Vigência: 12 meses. Valor mensal: R$ 3.000,00.",
    "Folha de pagamento dos colaboradores. Total de funcionários: 8. "
    "Total bruto: R$ 32.000,00. INSS: R$ 3.520,00.",
    "Declaração de Débitos e Créditos Tributários Federais. "
    "Competência: mês de referência. PIS/COFINS: R$ 1.800,00.",
    "Extrato bancário consolidado. Créditos: R$ 85.000,00. Débitos: R$ 72.000,00.",
    "Aditivo contratual alterando o prazo de vigência por mais 6 meses.",
    "Documento de rescisão de contrato de trabalho. Aviso prévio cumprido.",
    "Imagem digitalizada de documento original. Conteúdo verificado.",
    "Sped Fiscal — Registro C100. Valor das operações: R$ 45.000,00.",
]

DEMO_USER_USERNAME = "demo"
DEMO_USER_FULLNAME = "Usuário Demo"
# E-mail não é segredo (fixo, igual ao backend em app/config.py). A senha
# NUNCA fica hardcoded aqui — sempre lida de DEMO_PASSWORD (Fly secret).
# Incidente real: GitGuardian detectou a senha commitada quando ela estava
# direto como string neste arquivo e em SessionExpiredOverlay.tsx.
DEMO_USER_EMAIL = settings.DEMO_EMAIL
DEMO_USER_PASSWORD = settings.DEMO_PASSWORD

# Usuários de equipe extras — só existem pra popular Configurações → Usuários
# & Papéis com mais de uma linha (mostrando os 3 níveis de permissão em uso
# de verdade). Ninguém precisa logar com essas contas — a senha é gerada e
# descartada a cada reset (nunca exposta), só existe pra satisfazer a API do
# Supabase Auth, que exige uma senha pra criar a conta.
# (email, username, full_name, permission_level em user_company_access)
EXTRA_USERS = [
    ("operador.demo@docke.app", "operador.demo", "Operador Demo", "operador"),
    ("visualizador.demo@docke.app", "visualizador.demo", "Visualizadora Demo", "visualizador"),
]

ACTIONS = ["upload", "view", "download", "view", "view", "download"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def rand_date(days_back: int = 7) -> datetime:
    delta = timedelta(seconds=random.randint(0, days_back * 86400))
    return datetime.now(timezone.utc) - delta


def make_label() -> str:
    return "f" + str(int(datetime.now(timezone.utc).timestamp() * 1000)) + str(random.randint(1000, 9999))


# ---------------------------------------------------------------------------
# Geração de conteúdo real por tipo de arquivo (para os documentos poderem
# ser de fato abertos/baixados/pré-visualizados no modo demo — sem isso, o
# registro existe no banco mas o objeto nunca existiu no R2).
# ---------------------------------------------------------------------------

def _generate_minimal_pdf(text: str) -> bytes:
    """PDF de uma página, texto simples, construído manualmente (sem libs de PDF)."""
    safe_text = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    wrapped = [safe_text[i:i + 80] for i in range(0, len(safe_text), 80)] or [""]
    content_ops = " ".join(f"({line}) Tj 0 -14 Td" for line in wrapped)
    stream = f"BT /F1 11 Tf 50 720 Td {content_ops} ET"
    stream_bytes = stream.encode("latin-1", errors="replace")

    objects = [
        b"<</Type/Catalog/Pages 2 0 R>>",
        b"<</Type/Pages/Kids[3 0 R]/Count 1>>",
        b"<</Type/Page/Parent 2 0 R/Resources<</Font<</F1 4 0 R>>>>/MediaBox[0 0 612 792]/Contents 5 0 R>>",
        b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
        f"<</Length {len(stream_bytes)}>>stream\n".encode() + stream_bytes + b"\nendstream",
    ]

    out = io.BytesIO()
    out.write(b"%PDF-1.4\n")
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(out.tell())
        out.write(f"{i} 0 obj".encode())
        out.write(obj)
        out.write(b"endobj\n")
    xref_offset = out.tell()
    out.write(f"xref\n0 {len(objects) + 1}\n".encode())
    out.write(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.write(f"{off:010d} 00000 n \n".encode())
    out.write(f"trailer<</Size {len(objects) + 1}/Root 1 0 R>>\n".encode())
    out.write(f"startxref\n{xref_offset}\n%%EOF".encode())
    return out.getvalue()


def _generate_minimal_xlsx(title: str, text: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws["A1"] = title
    ws["A2"] = text
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_minimal_jpg(text: str) -> bytes:
    img = Image.new("RGB", (900, 700), color=(248, 248, 246))
    draw = ImageDraw.Draw(img)
    draw.multiline_text((30, 30), text, fill=(30, 30, 30))
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return buf.getvalue()


def _generate_minimal_docx(text: str) -> bytes:
    escaped = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        '</Relationships>'
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f'<w:body><w:p><w:r><w:t>{escaped}</w:t></w:r></w:p></w:body>'
        '</w:document>'
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document)
    return buf.getvalue()


def _generate_minimal_xml(title: str, text: str) -> bytes:
    escaped = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        f"<documento>\n  <titulo>{title}</titulo>\n  <conteudo>{escaped}</conteudo>\n</documento>\n"
    ).encode("utf-8")


def generate_file_content(ext: str, title: str, text: str) -> bytes:
    """Gera bytes de um arquivo real e abrível para o tipo dado (modo demo)."""
    if ext == "txt":
        return f"{title}\n\n{text}".encode("utf-8")
    if ext == "xml":
        return _generate_minimal_xml(title, text)
    if ext == "pdf":
        return _generate_minimal_pdf(f"{title}\n\n{text}")
    if ext == "xlsx":
        return _generate_minimal_xlsx(title, text)
    if ext == "jpg":
        return _generate_minimal_jpg(f"{title}\n\n{text}")
    if ext == "docx":
        return _generate_minimal_docx(f"{title}\n\n{text}")
    return text.encode("utf-8")


# ---------------------------------------------------------------------------
# Main seed
# ---------------------------------------------------------------------------

async def _ensure_auth_account(email: str, password: str, *, reset_password: bool = True) -> str:
    """
    Garante que existe uma conta real no Supabase Auth para o e-mail dado
    (mesmo padrão do ADR-033 em companies.py: Admin API, email_confirm=True,
    sem disparar e-mail de convite). Idempotente: se já existir, reaproveita
    o id em vez de recriar — evita acumular contas órfãs no Auth a cada reset.
    Generalizado a partir do antigo `_ensure_demo_auth_account` (agora usado
    também para os usuários de equipe extras — EXTRA_USERS).
    """
    async with httpx.AsyncClient() as client:
        # ATENÇÃO: a Admin API do GoTrue/Supabase NÃO filtra por e-mail via
        # query param — ela ignora silenciosamente qualquer parâmetro
        # desconhecido e devolve a listagem padrão. Filtrar do lado do
        # cliente é obrigatório aqui; usar existing[0] sem filtrar pegaria
        # o primeiro usuário da lista (ex: a conta admin real de produção),
        # não necessariamente o usuário demo. Isso já causou um incidente
        # real (senha e nome da conta admin real sobrescritos) — nunca
        # remover este filtro explícito por e-mail.
        list_resp = await client.get(
            f"{settings.SUPABASE_URL}/auth/v1/admin/users",
            headers={
                "apikey": settings.SUPABASE_SERVICE_ROLE_KEY,
                "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}",
            },
            params={"per_page": 200},
            timeout=10.0,
        )
        all_users = list_resp.json().get("users", []) if list_resp.status_code == 200 else []
        existing = [u for u in all_users if u.get("email", "").lower() == email.lower()]
        if existing:
            user_id = existing[0]["id"]
            if reset_password:
                await client.put(
                    f"{settings.SUPABASE_URL}/auth/v1/admin/users/{user_id}",
                    headers={
                        "apikey": settings.SUPABASE_SERVICE_ROLE_KEY,
                        "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}",
                        "Content-Type": "application/json",
                    },
                    json={"password": password},
                    timeout=10.0,
                )
            return user_id

        create_resp = await client.post(
            f"{settings.SUPABASE_URL}/auth/v1/admin/users",
            headers={
                "apikey": settings.SUPABASE_SERVICE_ROLE_KEY,
                "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY}",
                "Content-Type": "application/json",
            },
            json={"email": email, "password": password, "email_confirm": True},
            timeout=10.0,
        )
        if create_resp.status_code not in (200, 201):
            raise RuntimeError(f"Falha ao criar conta no Supabase Auth ({email}): {create_resp.text}")
        return create_resp.json()["id"]


def _hash_share_token(token: str) -> str:
    return hashlib.sha256(token.encode()).hexdigest()


async def _build_demo_data(
    conn: asyncpg.Connection,
    demo_user_id: str,
    extra_user_ids: dict[str, str],
) -> None:
    """
    Núcleo do seed — recebe uma conexão já aberta (própria ou do pool admin
    do app) e os ids das contas de Auth já garantidas. Chamado tanto pelo
    `run_seed()` (CLI) quanto por `demo_reset_service.reset_demo_data()`
    (worker periódico + endpoint de reset manual). Assume que o chamador já
    está dentro de uma transação.
    """
    # ------------------------------------------------------------------
    # 0. Limpar dados de demo anteriores (idempotente) — mantém os usuários
    #    (mesmos ids do Auth), só limpa as empresas fictícias.
    #    documents.company_id e activity_log.company_id são ON DELETE
    #    RESTRICT (não CASCADE) — apagar companies direto falha com FK
    #    violation assim que a 1ª rodada do seed já tiver criado dados
    #    reais. Apagar essas duas tabelas primeiro (o resto — folders,
    #    user_company_access, shares, notifications — já cascadeia de
    #    companies) resolve. documents também cascadeia pra ocr_jobs/
    #    favorites/document_versions/share_accesses (via shares).
    # ------------------------------------------------------------------
    old_company_ids = await conn.fetch(
        "SELECT id FROM public.companies WHERE name = ANY($1::text[])",
        [c[0] for c in COMPANIES],
    )
    old_ids = [r["id"] for r in old_company_ids]
    if old_ids:
        await conn.execute("DELETE FROM public.activity_log WHERE company_id = ANY($1::uuid[])", old_ids)
        await conn.execute("DELETE FROM public.documents WHERE company_id = ANY($1::uuid[])", old_ids)
        await conn.execute("DELETE FROM public.companies WHERE id = ANY($1::uuid[])", old_ids)
    print("✔ Dados anteriores removidos.")

    # ------------------------------------------------------------------
    # 1. Criar/atualizar usuário demo + usuários de equipe extras em
    #    public.users (mesmos ids das contas de Auth já garantidas)
    # ------------------------------------------------------------------
    await conn.execute(
        """
        INSERT INTO public.users (id, username, full_name, role)
        VALUES ($1::uuid, $2, $3, 'admin')
        ON CONFLICT (id) DO UPDATE SET username = EXCLUDED.username, full_name = EXCLUDED.full_name
        """,
        demo_user_id,
        DEMO_USER_USERNAME,
        DEMO_USER_FULLNAME,
    )
    print(f"✔ Usuário demo criado: {DEMO_USER_USERNAME} (id={demo_user_id[:8]}…)")

    for email, username, full_name, _permission in EXTRA_USERS:
        user_id = extra_user_ids[email]
        await conn.execute(
            """
            INSERT INTO public.users (id, username, full_name, role)
            VALUES ($1::uuid, $2, $3, 'usuario')
            ON CONFLICT (id) DO UPDATE SET username = EXCLUDED.username, full_name = EXCLUDED.full_name
            """,
            user_id,
            username,
            full_name,
        )
    print(f"✔ {len(EXTRA_USERS)} usuários de equipe extras criados.")

    for company_name, _ in COMPANIES:
        # ---------------------------------------------------------------
        # 2. Criar empresa
        # ---------------------------------------------------------------
        company_id = str(uuid.uuid4())
        await conn.execute(
            "INSERT INTO public.companies (id, name) VALUES ($1::uuid, $2)",
            company_id,
            company_name,
        )

        # ---------------------------------------------------------------
        # 3. Conceder acesso — demo como admin, equipe extra com o
        #    permission_level de cada um (visualizador/operador)
        # ---------------------------------------------------------------
        await conn.execute(
            """
            INSERT INTO public.user_company_access
                (user_id, company_id, permission_level, folder_path)
            VALUES ($1::uuid, $2::uuid, 'admin', NULL)
            """,
            demo_user_id,
            company_id,
        )
        for email, _username, _full_name, permission_level in EXTRA_USERS:
            await conn.execute(
                """
                INSERT INTO public.user_company_access
                    (user_id, company_id, permission_level, folder_path)
                VALUES ($1::uuid, $2::uuid, $3, NULL)
                """,
                extra_user_ids[email],
                company_id,
                permission_level,
            )

        # ---------------------------------------------------------------
        # 4. Criar pastas raiz
        # ---------------------------------------------------------------
        folder_ids: list[str] = []
        for folder_name in FOLDERS:
            folder_id = str(uuid.uuid4())
            label = make_label()
            await conn.execute(
                """
                INSERT INTO public.folders
                    (id, company_id, parent_id, path, name, created_by)
                VALUES ($1::uuid, $2::uuid, NULL, $3::ltree, $4, $5::uuid)
                """,
                folder_id,
                company_id,
                label,
                folder_name,
                demo_user_id,
            )
            folder_ids.append(folder_id)
            await asyncio.sleep(0.001)  # garante labels distintos

        # ---------------------------------------------------------------
        # 5. Criar documentos
        # ---------------------------------------------------------------
        docs_created: list[tuple[str, str]] = []  # (doc_id, name)

        for i in range(1, 18):  # ~17 docs por empresa → ~51 total
            tmpl, ext, folder_idx = random.choice(DOCUMENTS)
            doc_name = tmpl.format(n=str(i).zfill(2))
            folder_id = folder_ids[folder_idx]

            doc_id = str(uuid.uuid4())
            ocr_text = random.choice(OCR_TEXTS)

            # 2 docs com OCR failed, 3 na lixeira (para os últimos docs)
            if i >= 16:
                ocr_status = "failed"
                ocr_text = None
            elif i == 15:
                ocr_status = "failed"
                ocr_text = None
            else:
                ocr_status = "done"

            deleted_at = None
            deleted_original_folder_id = None
            if i >= 14 and i <= 16:
                # Coloca na lixeira (soft-delete) — só para os últimos 3
                deleted_at = rand_date(5)
                deleted_original_folder_id = folder_id

            mime_map = {
                "pdf": "application/pdf",
                "xml": "application/xml",
                "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "jpg": "image/jpeg",
                "txt": "text/plain",
            }
            mime_type = mime_map.get(ext, "application/octet-stream")

            created_at = rand_date(30)

            # Conteúdo real (não só o registro no banco) — sem isso o
            # arquivo nunca existiu no R2 e não pode ser aberto/baixado.
            storage_key = f"documents/{company_id}/{doc_id}.{ext}"
            file_bytes = generate_file_content(ext, doc_name, ocr_text or "Documento de exemplo do modo demo.")
            storage_service.put_object_bytes(storage_key, file_bytes, mime_type)
            size = len(file_bytes)
            content_hash = hashlib.sha256(file_bytes).hexdigest()

            await conn.execute(
                """
                INSERT INTO public.documents
                    (id, company_id, folder_id, name, mime_type, file_type, size_bytes,
                     storage_path, content_hash, ocr_status, ocr_text,
                     ocr_completed_at, deleted_at, deleted_original_folder_id,
                     uploaded_by, created_at)
                VALUES
                    ($1::uuid, $2::uuid, $3::uuid, $4, $5, $6, $7,
                     $8, $9, $10, $11,
                     $12, $13, $14,
                     $15::uuid, $16)
                """,
                doc_id,
                company_id,
                folder_id,
                doc_name,
                mime_type,
                ext,
                size,
                storage_key,
                content_hash,
                ocr_status,
                ocr_text,
                datetime.now(timezone.utc) if ocr_status == "done" else None,
                deleted_at,
                deleted_original_folder_id,
                demo_user_id,
                created_at,
            )

            # Versão 1 do histórico (ADR-024/029) — sem isso a aba "Versões"
            # fica vazia e um futuro "Enviar nova versão" seria rotulado
            # incorretamente como "Versão 1" (mesmo bug real corrigido em
            # confirm_upload_transaction, ver documents_service.py).
            version_id = await conn.fetchval(
                """
                INSERT INTO public.document_versions
                    (document_id, version_number, storage_key, size_bytes, mime_type, ocr_text, ocr_status, uploaded_by, created_at)
                VALUES
                    ($1::uuid, 1, $2, $3, $4, $5, $6, $7::uuid, $8)
                RETURNING id
                """,
                doc_id, storage_key, size, mime_type, ocr_text, ocr_status, demo_user_id, created_at,
            )
            await conn.execute(
                "UPDATE public.documents SET current_version_id = $2 WHERE id = $1",
                doc_id, version_id,
            )
            docs_created.append((doc_id, doc_name))

        # ---------------------------------------------------------------
        # 6. Atividade simulada (últimos 7 dias)
        # ---------------------------------------------------------------
        non_deleted_docs = [d for d in docs_created if docs_created.index(d) < 13]
        for _ in range(20):
            doc_id, doc_name = random.choice(non_deleted_docs)
            action = random.choice(ACTIONS)
            await conn.execute(
                """
                INSERT INTO public.activity_log
                    (user_id, company_id, action, item_type, item_id,
                     item_name_snapshot, created_at)
                VALUES ($1::uuid, $2::uuid, $3, 'document', $4::uuid, $5, $6)
                """,
                demo_user_id,
                company_id,
                action,
                doc_id,
                doc_name,
                rand_date(7),
            )

        # ---------------------------------------------------------------
        # 7. Favoritos (âncoras) — 2 documentos + 1 pasta, do usuário demo
        # ---------------------------------------------------------------
        anchor_docs = non_deleted_docs[:2]
        for doc_id, _doc_name in anchor_docs:
            await conn.execute(
                "INSERT INTO public.favorites (user_id, document_id) VALUES ($1::uuid, $2::uuid)",
                demo_user_id, doc_id,
            )
        await conn.execute(
            "INSERT INTO public.favorites (user_id, folder_id) VALUES ($1::uuid, $2::uuid)",
            demo_user_id, folder_ids[0],
        )

        # ---------------------------------------------------------------
        # 8. Links de compartilhamento — 1 aberto, 1 com senha
        # ---------------------------------------------------------------
        share_docs = non_deleted_docs[2:4]
        for idx, (doc_id, _doc_name) in enumerate(share_docs):
            token = uuid.uuid4().hex + uuid.uuid4().hex
            token_hash = _hash_share_token(token)
            password_hash = (
                bcrypt.hashpw(b"demo1234", bcrypt.gensalt()).decode() if idx == 1 else None
            )
            await conn.execute(
                """
                INSERT INTO public.shares
                    (resource_type, resource_id, company_id, token_hash, password_hash, expires_at, created_by)
                VALUES ('document', $1::uuid, $2::uuid, $3, $4, now() + interval '30 days', $5::uuid)
                """,
                doc_id, company_id, token_hash, password_hash, demo_user_id,
            )

        print(f"  ✔ {company_name}: {len(docs_created)} docs, {len(folder_ids)} pastas, "
              f"{len(anchor_docs) + 1} favoritos, {len(share_docs)} links")


async def run_seed() -> None:
    """CLI: `python -m app.seed.demo_data` — abre sua própria conexão."""
    if not DEMO_USER_PASSWORD:
        raise RuntimeError(
            "DEMO_PASSWORD não está configurada (Fly secret). "
            "Rode: fly secrets set DEMO_PASSWORD='...' antes de rodar o seed."
        )
    demo_user_id = await _ensure_auth_account(DEMO_USER_EMAIL, DEMO_USER_PASSWORD)
    print(f"✔ Conta demo no Supabase Auth pronta (id={demo_user_id[:8]}…).")

    extra_user_ids: dict[str, str] = {}
    for email, _username, _full_name, _permission in EXTRA_USERS:
        extra_user_ids[email] = await _ensure_auth_account(email, secrets.token_urlsafe(24))
    print(f"✔ {len(EXTRA_USERS)} contas de equipe extras no Supabase Auth prontas.")

    conn = await asyncpg.connect(settings.asyncpg_url)
    print("✔ Conectado ao banco de dados.")

    try:
        await conn.execute("BEGIN")
        await _build_demo_data(conn, demo_user_id, extra_user_ids)
        await conn.execute("COMMIT")
        print("\n✅ Seed concluído com sucesso!")
        print(f"   Login demo: {DEMO_USER_EMAIL} / {DEMO_USER_PASSWORD}")
        print("   Empresas: Comércio Alfa Modelo Ltda | Serviços Beta Referência SA | Indústria Gama Exemplo Ltda")
    except Exception as exc:
        await conn.execute("ROLLBACK")
        print(f"\n❌ Erro durante o seed: {exc}")
        raise
    finally:
        await conn.close()


if __name__ == "__main__":
    asyncio.run(run_seed())
